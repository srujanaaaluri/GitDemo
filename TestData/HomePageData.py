import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname" : "rahul","lastname":"shetty","password":"jvfgvv"}, {"firstname":"suma","lastname": "shetty","password":"ygy"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\sruja\\Documents\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(1, sheet.max_column + 1):  # to get columns
                    # dict[
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[(Dict)]